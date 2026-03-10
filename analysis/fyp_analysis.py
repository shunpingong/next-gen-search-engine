from __future__ import annotations

from pathlib import Path
from typing import Dict, Tuple

import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
from openpyxl import load_workbook


# Colors used consistently across figures for easier comparison.
CONFIG_COLOR_MAP = {
    "LLM-only": "#4C78A8",
    "RAG": "#F58518",
    "TAG": "#54A24B",
    "Unknown": "#B279A2",
}

TOOL_COLOR_MAP = {
    "None": "#9E9E9E",
    "Basic": "#1F77B4",
    "Detailed": "#D62728",
}

QUERY_TYPE_COLOR_MAP = {
    "Static": "#2A9D8F",
    "Dynamic": "#E76F51",
}


def set_academic_plot_style() -> None:
    """Set a clean seaborn + matplotlib style suitable for academic reports."""
    sns.set_theme(style="whitegrid", context="paper")
    plt.rcParams.update(
        {
            "figure.dpi": 150,
            "savefig.dpi": 300,
            "axes.titlesize": 12,
            "axes.labelsize": 10,
            "xtick.labelsize": 9,
            "ytick.labelsize": 9,
            "legend.fontsize": 9,
        }
    )


def _build_palette(
    categories: list[str],
    preferred_map: dict[str, str] | None = None,
    fallback_palette: str = "tab10",
) -> dict[str, str]:
    """Build a deterministic palette mapping for categorical values."""
    values = [str(c) for c in pd.Series(categories).dropna().unique().tolist()]
    mapping: dict[str, str] = {}
    if preferred_map:
        mapping.update({k: v for k, v in preferred_map.items() if k in values})

    missing = [c for c in values if c not in mapping]
    if missing:
        fallback_colors = sns.color_palette(fallback_palette, n_colors=len(missing))
        for category, color in zip(missing, fallback_colors):
            mapping[category] = color
    return mapping


def _model_palette(series: pd.Series) -> dict[str, str]:
    """Color mapping for model categories."""
    models = sorted(series.dropna().astype(str).unique().tolist())
    return _build_palette(models, preferred_map=None, fallback_palette="tab20")


def _save_and_close(fig: plt.Figure, output_dir: Path, filename: str) -> None:
    output_path = output_dir / filename
    fig.tight_layout()
    fig.savefig(output_path, bbox_inches="tight")
    plt.close(fig)


def _to_percent_if_ratio(series: pd.Series) -> pd.Series:
    """Convert 0-1 ratios to percentages; keep existing percentage values unchanged."""
    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.dropna().empty:
        return numeric
    if float(numeric.max(skipna=True)) <= 1.0 + 1e-9:
        return numeric * 100
    return numeric


def _coherence_to_percent(series: pd.Series, max_score: float = 5.0) -> pd.Series:
    """Convert coherence score to percentage scale using max_score."""
    numeric = pd.to_numeric(series, errors="coerce")
    return (numeric / max_score) * 100


def resolve_input_path(file_name: str, output_dir: Path) -> Path:
    """
    Resolve dataset path.
    Tries provided file_name first, then falls back to 'FYP Results.xlsx'.
    """
    candidate = output_dir / file_name
    if candidate.exists():
        return candidate

    fallback = output_dir / "FYP Results.xlsx"
    if fallback.exists():
        return fallback

    raise FileNotFoundError(
        f"Could not find '{file_name}' or 'FYP Results.xlsx' inside {output_dir}."
    )


def _normalize_configuration_value(value: object) -> str:
    """Normalize configuration labels to canonical names."""
    if pd.isna(value):
        return "Unknown"

    raw = str(value).strip()
    key = raw.upper()
    normalized_map = {
        "C1": "LLM-only",
        "C2": "RAG",
        "C3": "TAG",
        "LLM": "LLM-only",
        "LLM-ONLY": "LLM-only",
        "LLM ONLY": "LLM-only",
        "LLM + RAG": "RAG",
        "LLM+RAG": "RAG",
        "RAG": "RAG",
        "LLM + TAG": "TAG",
        "LLM+TAG": "TAG",
        "TAG": "TAG",
    }
    return normalized_map.get(key, raw)


def _extract_main_table_from_raw_sheet(
    raw_sheet: pd.DataFrame,
    forced_header_row_idx: int | None = None,
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """
    Extract the query-level table from sheet 1 that has metadata above the real header.
    Returns (main_table_df, config_code_map).
    """
    required_header_tokens = {"Query_ID", "Query_Type", "Query_Text", "Model_Name", "Configuration"}
    header_row_idx = forced_header_row_idx

    if header_row_idx is None:
        for idx in raw_sheet.index:
            row_tokens = set(raw_sheet.loc[idx].dropna().astype(str).str.strip().tolist())
            if required_header_tokens.issubset(row_tokens):
                header_row_idx = idx
                break

    if header_row_idx is None:
        raise ValueError(
            "Could not locate query-level header row in Sheet 1. "
            "Expected columns include Query_ID, Query_Type, Query_Text, Model_Name, Configuration."
        )

    config_code_map: Dict[str, str] = {}
    for idx in raw_sheet.index:
        if idx >= header_row_idx:
            break
        config_value = raw_sheet.iat[idx, 3] if raw_sheet.shape[1] > 3 else None
        config_code = raw_sheet.iat[idx, 4] if raw_sheet.shape[1] > 4 else None
        if pd.notna(config_code) and pd.notna(config_value):
            code_text = str(config_code).strip()
            if code_text.upper().startswith("C") and len(code_text) <= 4:
                config_code_map[code_text.upper()] = _normalize_configuration_value(config_value)

    headers = [str(x).strip() for x in raw_sheet.loc[header_row_idx].tolist()]
    main_df = raw_sheet.loc[header_row_idx + 1 :].copy()
    main_df.columns = headers

    main_df = main_df.loc[:, [c for c in main_df.columns if not str(c).startswith("Unnamed")]]
    main_df = main_df.dropna(how="all").reset_index(drop=True)

    # Only enforce Query_ID filtering if Query_ID values exist in this workbook.
    if "Query_ID" in main_df.columns and main_df["Query_ID"].notna().any():
        main_df = main_df[main_df["Query_ID"].notna()].reset_index(drop=True)

    return main_df, config_code_map


def load_main_sheet(file_path: Path, main_header_row_excel: int = 22) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Load the main query-level sheet and config code map from workbook."""
    raw_sheet = pd.read_excel(file_path, sheet_name=0, header=None)
    forced_header_row_idx = max(main_header_row_excel - 1, 0)
    main_df, config_code_map = _extract_main_table_from_raw_sheet(
        raw_sheet, forced_header_row_idx=forced_header_row_idx
    )
    return main_df, config_code_map


def clean_main_dataset(df: pd.DataFrame, config_code_map: Dict[str, str] | None = None) -> pd.DataFrame:
    """Clean and standardize the main experiment dataset."""
    cleaned = df.copy()
    cleaned.columns = [str(col).strip() for col in cleaned.columns]

    aliases = {
        "Model": "Model_Name",
        "Latency_Total": "Latency_Total (s)",
        "Retrieval_Time": "Retrieval_Time (s)",
        "Ranking_Time": "Ranking_Time (s)",
    }
    cleaned = cleaned.rename(columns={k: v for k, v in aliases.items() if k in cleaned.columns})

    text_cols = [
        "Query_ID",
        "Query_Type",
        "Model_Name",
        "Configuration",
        "Tool_Invoked",
        "Query_Refined",
        "TextGrad_Applied",
        "PageRank_Applied",
    ]
    for col in text_cols:
        if col in cleaned.columns:
            cleaned[col] = cleaned[col].astype("string").str.strip()

    numeric_cols = [
        "Factual_Accuracy",
        "Coherence_Score",
        "Latency_Total (s)",
        "Retrieval_Time (s)",
        "Ranking_Time (s)",
        "Num_Documents_Retrieved",
    ]
    for col in numeric_cols:
        if col in cleaned.columns:
            cleaned[col] = pd.to_numeric(cleaned[col], errors="coerce")

    if "Configuration" in cleaned.columns:
        if config_code_map:
            code_map = {str(k).strip().upper(): _normalize_configuration_value(v) for k, v in config_code_map.items()}
        else:
            code_map = {}
        cleaned["Configuration"] = cleaned["Configuration"].apply(
            lambda x: code_map.get(str(x).strip().upper(), _normalize_configuration_value(x))
        )

    if "Query_Type" in cleaned.columns:
        cleaned["Query_Type"] = cleaned["Query_Type"].replace(
            {"static": "Static", "dynamic": "Dynamic", "STATIC": "Static", "DYNAMIC": "Dynamic"}
        )

    if "Tool_Invoked" in cleaned.columns:
        cleaned["Tool_Invoked"] = cleaned["Tool_Invoked"].fillna("None")
        cleaned["Tool_Invoked"] = cleaned["Tool_Invoked"].replace(
            {
                "none": "None",
                "basic": "Basic",
                "detailed": "Detailed",
                "<NA>": "None",
            }
        )

    required_cols = [
        "Model_Name",
        "Configuration",
        "Query_Type",
        "Factual_Accuracy",
        "Coherence_Score",
        "Latency_Total (s)",
    ]
    missing_required = [c for c in required_cols if c not in cleaned.columns]
    if missing_required:
        raise ValueError(f"Missing required columns in main dataset after cleaning: {missing_required}")

    cleaned = cleaned.dropna(subset=required_cols).reset_index(drop=True)
    return cleaned


def reset_workbook_to_main_sheet(workbook_path: Path, main_sheet_name: str | None = None) -> str:
    """
    Keep only the main sheet in the workbook and delete all others.
    Returns the kept main sheet name.
    """
    wb = load_workbook(workbook_path)
    if not wb.sheetnames:
        wb.close()
        raise ValueError(f"No sheets found in workbook: {workbook_path}")

    keep_sheet = main_sheet_name or wb.sheetnames[0]
    if keep_sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Main sheet '{keep_sheet}' not found in workbook.")

    for sheet in list(wb.sheetnames):
        if sheet != keep_sheet:
            del wb[sheet]

    wb.save(workbook_path)
    wb.close()
    return keep_sheet


def write_tables_to_existing_workbook(
    workbook_path: Path,
    tables: Dict[str, pd.DataFrame],
    main_sheet_name: str | None = None,
) -> None:
    """
    Reset workbook to keep only main sheet, then write fresh summary tables as new sheets.
    """
    if workbook_path.exists():
        reset_workbook_to_main_sheet(workbook_path, main_sheet_name=main_sheet_name)
        mode = "a"
    else:
        mode = "w"

    writer_kwargs = {"engine": "openpyxl", "mode": mode}
    if mode == "a":
        writer_kwargs["if_sheet_exists"] = "replace"

    with pd.ExcelWriter(workbook_path, **writer_kwargs) as writer:
        for sheet_name, table_df in tables.items():
            table_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)


def section_5_1_overall(df: pd.DataFrame, output_dir: Path) -> Dict[str, pd.DataFrame]:
    """
    Figure 28: Accuracy across models and configurations.
    Figure 29: Latency by configuration.
    """
    model_config = (
        df.groupby(["Model_Name", "Configuration"], as_index=False)
        .agg(Avg_Accuracy=("Factual_Accuracy", "mean"), Avg_Latency=("Latency_Total (s)", "mean"))
        .sort_values(["Model_Name", "Configuration"])
    )
    model_config["Avg_Accuracy (%)"] = _to_percent_if_ratio(model_config["Avg_Accuracy"])

    config_summary = (
        df.groupby("Configuration", as_index=False)
        .agg(
            Avg_Accuracy=("Factual_Accuracy", "mean"),
            Avg_Latency=("Latency_Total (s)", "mean"),
        )
        .sort_values("Configuration")
    )
    config_summary["Avg_Accuracy (%)"] = _to_percent_if_ratio(config_summary["Avg_Accuracy"])

    config_palette = _build_palette(
        model_config["Configuration"].dropna().astype(str).tolist(),
        preferred_map=CONFIG_COLOR_MAP,
        fallback_palette="Set2",
    )

    fig, ax = plt.subplots(figsize=(10, 5))
    sns.barplot(
        data=model_config,
        x="Model_Name",
        y="Avg_Accuracy (%)",
        hue="Configuration",
        palette=config_palette,
        ax=ax,
    )
    ax.set_title("Model Accuracy Comparison Across Configurations")
    ax.set_ylabel("Average Accuracy (%)")
    ax.set_xlabel("Model")
    ax.set_ylim(0, 100)
    _save_and_close(fig, output_dir, "fig_28_accuracy_model_config.png")

    fig, ax = plt.subplots(figsize=(8, 5))
    sns.barplot(
        data=config_summary,
        x="Configuration",
        y="Avg_Latency",
        hue="Configuration",
        palette=config_palette,
        legend=False,
        ax=ax,
    )
    ax.set_title("Average System Latency by Configuration")
    ax.set_ylabel("Average Latency (s)")
    ax.set_xlabel("Configuration")
    _save_and_close(fig, output_dir, "fig_29_latency_by_configuration.png")

    return {
        "model_config": model_config,
        "config_summary": config_summary,
    }


def section_5_2_static_dynamic(df: pd.DataFrame, output_dir: Path) -> pd.DataFrame:
    """
    Figure 30: Static vs Dynamic accuracy (grouped by configuration).
    Figure 31: Static vs Dynamic latency (grouped by configuration).
    """
    agg = (
        df.groupby(["Query_Type", "Configuration"], as_index=False)
        .agg(
            Avg_Accuracy=("Factual_Accuracy", "mean"),
            Avg_Latency=("Latency_Total (s)", "mean"),
        )
        .sort_values(["Query_Type", "Configuration"])
    )
    agg["Avg_Accuracy (%)"] = _to_percent_if_ratio(agg["Avg_Accuracy"])

    config_palette = _build_palette(
        agg["Configuration"].dropna().astype(str).tolist(),
        preferred_map=CONFIG_COLOR_MAP,
        fallback_palette="Set2",
    )

    fig, ax = plt.subplots(figsize=(8, 5))
    sns.barplot(
        data=agg,
        x="Query_Type",
        y="Avg_Accuracy (%)",
        hue="Configuration",
        palette=config_palette,
        ax=ax,
    )
    ax.set_title("Static vs Dynamic Accuracy")
    ax.set_ylabel("Average Accuracy (%)")
    ax.set_xlabel("Query Type")
    ax.set_ylim(0, 100)
    _save_and_close(fig, output_dir, "fig_30_static_dynamic_accuracy.png")

    fig, ax = plt.subplots(figsize=(8, 5))
    sns.barplot(
        data=agg,
        x="Query_Type",
        y="Avg_Latency",
        hue="Configuration",
        palette=config_palette,
        ax=ax,
    )
    ax.set_title("Static vs Dynamic Latency")
    ax.set_ylabel("Average Latency (s)")
    ax.set_xlabel("Query Type")
    _save_and_close(fig, output_dir, "fig_31_static_dynamic_latency.png")

    return agg


def section_5_3_tool_usage_tag(df: pd.DataFrame, output_dir: Path) -> Dict[str, pd.DataFrame]:
    """
    Figure 32: Tool invocation distribution (TAG only).
    Figure 33: Tool usage rate by model (TAG only).
    """
    tag_df = df[df["Configuration"].astype(str).str.upper() == "TAG"].copy()
    if tag_df.empty:
        empty_dist = pd.DataFrame(columns=["Tool_Invoked", "Count", "Percentage"])
        empty_rate = pd.DataFrame(columns=["Model_Name", "Tool_Usage_Rate", "Tool_Usage_Rate (%)"])
        return {"tool_distribution": empty_dist, "tool_usage_rate_by_model": empty_rate}

    tool_dist = (
        tag_df["Tool_Invoked"]
        .fillna("None")
        .value_counts(dropna=False)
        .rename_axis("Tool_Invoked")
        .reset_index(name="Count")
    )
    tool_dist["Percentage"] = 100 * tool_dist["Count"] / tool_dist["Count"].sum()
    tool_palette = _build_palette(
        tool_dist["Tool_Invoked"].astype(str).tolist(),
        preferred_map=TOOL_COLOR_MAP,
        fallback_palette="Set2",
    )

    fig, ax = plt.subplots(figsize=(7, 7))
    ax.pie(
        tool_dist["Count"],
        labels=tool_dist["Tool_Invoked"],
        autopct="%1.1f%%",
        startangle=90,
        colors=[tool_palette.get(str(v), "#999999") for v in tool_dist["Tool_Invoked"]],
    )
    ax.set_title("Distribution of Tool Invocation Types (TAG)")
    _save_and_close(fig, output_dir, "fig_32_tool_invocation_distribution.png")

    usage_df = tag_df.copy()
    usage_df["Tool_Used"] = (
        usage_df["Tool_Invoked"].fillna("None").astype(str).str.strip().str.lower() != "none"
    ).astype(int)
    tool_rate_by_model = (
        usage_df.groupby("Model_Name", as_index=False)["Tool_Used"]
        .mean()
        .rename(columns={"Tool_Used": "Tool_Usage_Rate"})
        .sort_values("Model_Name")
    )
    tool_rate_by_model["Tool_Usage_Rate (%)"] = _to_percent_if_ratio(tool_rate_by_model["Tool_Usage_Rate"])
    model_palette = _model_palette(tool_rate_by_model["Model_Name"])

    fig, ax = plt.subplots(figsize=(9, 5))
    sns.barplot(
        data=tool_rate_by_model,
        x="Model_Name",
        y="Tool_Usage_Rate (%)",
        hue="Model_Name",
        palette=model_palette,
        legend=False,
        ax=ax,
    )
    ax.set_title("Tool Usage Rate by Model (TAG)")
    ax.set_ylabel("Tool Usage Rate (%)")
    ax.set_xlabel("Model")
    ax.set_ylim(0, 100)
    _save_and_close(fig, output_dir, "fig_33_tool_usage_rate_by_model.png")

    return {
        "tool_distribution": tool_dist,
        "tool_usage_rate_by_model": tool_rate_by_model,
    }


def section_5_4_model_comparison(df: pd.DataFrame, output_dir: Path) -> pd.DataFrame:
    """
    Figure 34: Coherence comparison across models.
    Figure 35: Latency comparison across models.
    """
    model_agg = (
        df.groupby("Model_Name", as_index=False)
        .agg(
            Avg_Coherence=("Coherence_Score", "mean"),
            Avg_Latency=("Latency_Total (s)", "mean"),
        )
        .sort_values("Model_Name")
    )
    model_agg["Avg_Coherence (%)"] = _coherence_to_percent(model_agg["Avg_Coherence"])
    model_palette = _model_palette(model_agg["Model_Name"])

    fig, ax = plt.subplots(figsize=(8, 5))
    sns.barplot(
        data=model_agg,
        x="Model_Name",
        y="Avg_Coherence (%)",
        hue="Model_Name",
        palette=model_palette,
        legend=False,
        ax=ax,
    )
    ax.set_title("Coherence Comparison Across Models")
    ax.set_ylabel("Average Coherence (%)")
    ax.set_xlabel("Model")
    ax.set_ylim(0, 100)
    _save_and_close(fig, output_dir, "fig_34_model_coherence.png")

    fig, ax = plt.subplots(figsize=(8, 5))
    sns.barplot(
        data=model_agg,
        x="Model_Name",
        y="Avg_Latency",
        hue="Model_Name",
        palette=model_palette,
        legend=False,
        ax=ax,
    )
    ax.set_title("Latency Comparison Across Models")
    ax.set_ylabel("Average Latency (s)")
    ax.set_xlabel("Model")
    _save_and_close(fig, output_dir, "fig_35_model_latency.png")

    return model_agg


def section_5_5_error_analysis(df: pd.DataFrame, output_dir: Path) -> Dict[str, pd.DataFrame]:
    """
    Figure 36: Error rate comparison across models and configurations.
    Also returns configuration accuracy summary (requested instead of configuration error table).
    """
    working = df.copy()
    working["Is_Error"] = (working["Factual_Accuracy"] == 0).astype(int)

    error_model_config = (
        working.groupby(["Model_Name", "Configuration"], as_index=False)["Is_Error"]
        .mean()
        .rename(columns={"Is_Error": "Error_Rate"})
        .sort_values(["Model_Name", "Configuration"])
    )
    error_model_config["Error_Rate (%)"] = _to_percent_if_ratio(error_model_config["Error_Rate"])

    config_accuracy = (
        working.groupby("Configuration", as_index=False)["Factual_Accuracy"]
        .mean()
        .rename(columns={"Factual_Accuracy": "Accuracy"})
        .sort_values("Configuration")
    )
    config_accuracy["Accuracy (%)"] = _to_percent_if_ratio(config_accuracy["Accuracy"])

    config_palette = _build_palette(
        error_model_config["Configuration"].dropna().astype(str).tolist(),
        preferred_map=CONFIG_COLOR_MAP,
        fallback_palette="Set2",
    )

    fig, ax = plt.subplots(figsize=(10, 5))
    sns.barplot(
        data=error_model_config,
        x="Model_Name",
        y="Error_Rate (%)",
        hue="Configuration",
        palette=config_palette,
        ax=ax,
    )
    ax.set_title("Error Rate Comparison Across Models and Configurations")
    ax.set_ylabel("Error Rate (%)")
    ax.set_xlabel("Model")
    ax.set_ylim(0, 100)
    _save_and_close(fig, output_dir, "fig_36_error_rate_model_configuration.png")

    return {
        "error_rate_model_configuration": error_model_config,
        "configuration_accuracy": config_accuracy,
    }


def main(
    file_name: str = "results_dataset.xlsx",
    output_dir: str | Path = ".",
    main_header_row_excel: int = 22,
    main_sheet_name: str | None = None,
) -> Dict[str, pd.DataFrame]:
    """
    Run requested analyses (Figures 28-36), save figures, and refresh workbook sheets.
    """
    output_path = Path(output_dir)
    input_path = resolve_input_path(file_name=file_name, output_dir=output_path)

    set_academic_plot_style()
    main_raw, config_code_map = load_main_sheet(input_path, main_header_row_excel=main_header_row_excel)
    main_df = clean_main_dataset(main_raw, config_code_map=config_code_map)

    section_51 = section_5_1_overall(main_df, output_path)
    section_52 = section_5_2_static_dynamic(main_df, output_path)
    section_53 = section_5_3_tool_usage_tag(main_df, output_path)
    section_54 = section_5_4_model_comparison(main_df, output_path)
    section_55 = section_5_5_error_analysis(main_df, output_path)

    tables_to_write: Dict[str, pd.DataFrame] = {
        "S5_1_ModelConfig": section_51["model_config"],
        "S5_1_ConfigSummary": section_51["config_summary"],
        "S5_2_StaticDynamic": section_52,
        "S5_3_ToolDist": section_53["tool_distribution"],
        "S5_3_ToolRateByModel": section_53["tool_usage_rate_by_model"],
        "S5_4_ModelComparison": section_54,
        "S5_5_ErrorModelConfig": section_55["error_rate_model_configuration"],
        "S5_5_ConfigAccuracy": section_55["configuration_accuracy"],
    }
    write_tables_to_existing_workbook(input_path, tables_to_write, main_sheet_name=main_sheet_name)

    return {
        "section_5_1": section_51["model_config"],
        "section_5_1_config_summary": section_51["config_summary"],
        "section_5_2": section_52,
        "section_5_3_tool_distribution": section_53["tool_distribution"],
        "section_5_3_tool_rate_by_model": section_53["tool_usage_rate_by_model"],
        "section_5_4": section_54,
        "section_5_5_error_model_configuration": section_55["error_rate_model_configuration"],
        "section_5_5_configuration_accuracy": section_55["configuration_accuracy"],
    }


if __name__ == "__main__":
    # Running directly from analysis folder:
    # python fyp_analysis.py
    main(
        file_name="results_dataset.xlsx",
        output_dir=Path(__file__).resolve().parent,
        main_header_row_excel=22,
        main_sheet_name=None,
    )
